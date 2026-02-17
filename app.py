from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
from datetime import datetime
from io import BytesIO
import uuid

app = Flask(__name__)
CORS(app)

# ============================================================
# IN-MEMORY STORAGE (Works on Vercel - data resets on restart)
# For permanent storage, connect Supabase later
# ============================================================
TRIPS = []
EXPENSES = []

# ============================================================
# SERVE FRONTEND - Embedded HTML
# ============================================================
HTML_CONTENT = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Trip Expense Manager</title>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700;800&family=DM+Serif+Display&display=swap" rel="stylesheet">
    <style>
        :root {
            --primary: #6366f1;
            --primary-dark: #4f46e5;
            --secondary: #ec4899;
            --success: #10b981;
            --warning: #f59e0b;
            --danger: #ef4444;
            --dark: #1e293b;
            --light: #f8fafc;
            --gray: #64748b;
            --border: #e2e8f0;
        }

        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: 'Poppins', sans-serif;
            background-image: 
        linear-gradient(rgba(102, 126, 234, 0.3), rgba(118, 75, 162, 0.3)),
        url('https://images.unsplash.com/photo-1488646953014-85cb44e25828?w=1920&q=80');
    background-size: cover;
    background-position: center;
    background-attachment: fixed;
            min-height: 100vh;
            padding: 20px;
            color: var(--dark);
        }

        .container {
            max-width: 1400px;
            margin: 0 auto;
        }

        /* Header */
        .header {
            text-align: center;
            margin-bottom: 40px;
            animation: fadeInDown 0.8s ease;
        }

        .header h1 {
            font-family: 'DM Serif Display', serif;
            font-size: 4em;
            color: white;
            text-shadow: 0 4px 20px rgba(0,0,0,0.2);
            margin-bottom: 10px;
        }

        .header p {
            color: rgba(255,255,255,0.9);
            font-size: 1.2em;
            font-weight: 300;
        }

        /* Main Layout */
        .main-layout {
            display: grid;
            grid-template-columns: 350px 1fr;
            gap: 25px;
            animation: fadeInUp 0.8s ease 0.2s both;
        }

        /* Sidebar */
        .sidebar {
            background: white;
            border-radius: 20px;
            padding: 25px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.15);
            height: fit-content;
            position: sticky;
            top: 20px;
        }

        .sidebar h2 {
            font-size: 1.5em;
            margin-bottom: 20px;
            color: var(--dark);
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .new-trip-btn {
            width: 100%;
            padding: 15px;
            background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%);
            color: white;
            border: none;
            border-radius: 12px;
            font-size: 1em;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
        }

        .new-trip-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(99, 102, 241, 0.4);
        }

        .trip-list {
            max-height: 500px;
            overflow-y: auto;
        }

        .trip-card {
            background: var(--light);
            border: 2px solid var(--border);
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 12px;
            cursor: pointer;
            transition: all 0.3s;
            position: relative;
        }

        .trip-card:hover {
            border-color: var(--primary);
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(99, 102, 241, 0.2);
        }

        .trip-card.active {
            background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%);
            color: white;
            border-color: transparent;
        }

        .trip-card.completed {
            opacity: 0.7;
        }

        .trip-card h3 {
            font-size: 1.1em;
            margin-bottom: 8px;
        }

        .trip-card .budget {
            font-size: 0.9em;
            opacity: 0.8;
        }

        .trip-card .status-badge {
            position: absolute;
            top: 10px;
            right: 10px;
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 0.7em;
            font-weight: 600;
            background: rgba(255,255,255,0.3);
        }

        .trip-card.active .status-badge {
            background: rgba(255,255,255,0.2);
        }

        .empty-trips {
            text-align: center;
            padding: 40px 20px;
            color: var(--gray);
        }

        /* Main Content */
        .main-content {
            background: white;
            border-radius: 20px;
            padding: 35px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.15);
            min-height: 600px;
        }

        .trip-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 2px solid var(--border);
        }

        .trip-header-left {
            display: flex;
            align-items: center;
            gap: 15px;
        }

        .trip-header h2 {
            font-family: 'DM Serif Display', serif;
            font-size: 2.5em;
            color: var(--dark);
        }

        .edit-trip-btn {
            background: var(--gray);
            color: white;
            border: none;
            padding: 8px 15px;
            border-radius: 8px;
            cursor: pointer;
            font-size: 0.85em;
            transition: all 0.3s;
        }

        .edit-trip-btn:hover {
            background: var(--dark);
        }

        .trip-actions {
            display: flex;
            gap: 10px;
        }

        .btn {
            padding: 10px 20px;
            border: none;
            border-radius: 10px;
            font-size: 0.95em;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }

        .btn-complete {
            background: var(--success);
            color: white;
        }

        .btn-complete.reopening {
            background: var(--warning);
        }

        .btn-complete.reopening::before {
            content: '↻ ';
        }

        .btn-export {
            background: var(--warning);
            color: white;
        }

        .btn-delete {
            background: var(--danger);
            color: white;
        }

        .btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(0,0,0,0.15);
        }

        /* Budget Overview */
        .budget-overview {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 35px;
        }

        .budget-card {
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            padding: 25px;
            border-radius: 15px;
            color: white;
            position: relative;
            overflow: hidden;
        }

        .budget-card::before {
            content: '';
            position: absolute;
            top: -50%;
            right: -50%;
            width: 200%;
            height: 200%;
            background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
        }

        .budget-card.budget {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }

        .budget-card.spent {
            background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
        }

        .budget-card.remaining {
            background: linear-gradient(135deg, #30cfd0 0%, #330867 100%);
        }

        .budget-card h3 {
            font-size: 0.9em;
            opacity: 0.9;
            margin-bottom: 10px;
            font-weight: 400;
            position: relative;
        }

        .budget-card .amount {
            font-size: 2.5em;
            font-weight: 800;
            position: relative;
        }

        .budget-card .small-text {
            font-size: 0.8em;
            opacity: 0.8;
            margin-top: 5px;
        }

        /* Add Expense Form */
        .add-expense-section {
            background: var(--light);
            padding: 25px;
            border-radius: 15px;
            margin-bottom: 30px;
        }

        .add-expense-section h3 {
            margin-bottom: 20px;
            color: var(--dark);
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .form-grid {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr 1.5fr auto;
            gap: 15px;
            align-items: end;
        }

        .form-group {
            display: flex;
            flex-direction: column;
        }

        .form-group label {
            font-size: 0.85em;
            font-weight: 600;
            color: var(--gray);
            margin-bottom: 6px;
        }

        .form-group input,
        .form-group select {
            padding: 12px 15px;
            border: 2px solid var(--border);
            border-radius: 10px;
            font-size: 1em;
            font-family: 'Poppins', sans-serif;
            transition: all 0.3s;
        }

        .form-group input[type="file"] {
            padding: 8px 15px;
        }

        .image-preview {
            margin-top: 10px;
            max-width: 200px;
            max-height: 150px;
            border-radius: 8px;
            border: 2px solid var(--border);
            display: none;
        }

        .image-preview.active {
            display: block;
        }

        .camera-btn {
            background: var(--primary);
            color: white;
            border: none;
            padding: 10px 15px;
            border-radius: 8px;
            cursor: pointer;
            font-size: 0.9em;
            margin-top: 5px;
            transition: all 0.3s;
        }

        .camera-btn:hover {
            background: var(--primary-dark);
        }

        .form-group input:focus,
        .form-group select:focus {
            outline: none;
            border-color: var(--primary);
            box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.1);
        }

        .btn-add {
            background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%);
            color: white;
            padding: 12px 25px;
            border: none;
            border-radius: 10px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
            height: 48px;
        }

        .btn-add:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(99, 102, 241, 0.4);
        }

        /* Expense Table */
        .expense-table-container {
            margin-bottom: 20px;
        }

        .expense-table-container h3 {
            margin-bottom: 15px;
            color: var(--dark);
        }

        table {
            width: 100%;
            border-collapse: collapse;
        }

        thead {
            background: var(--dark);
            color: white;
        }

        th {
            padding: 15px;
            text-align: left;
            font-weight: 600;
            font-size: 0.9em;
        }

        td {
            padding: 15px;
            border-bottom: 1px solid var(--border);
        }

        tbody tr {
            transition: all 0.3s;
        }

        tbody tr:hover {
            background: var(--light);
        }

        .delete-expense-btn {
            background: var(--danger);
            color: white;
            border: none;
            padding: 6px 12px;
            border-radius: 6px;
            cursor: pointer;
            font-size: 0.85em;
            transition: all 0.3s;
        }

        .delete-expense-btn:hover {
            background: #dc2626;
        }

        .empty-state {
            text-align: center;
            padding: 60px 20px;
            color: var(--gray);
        }

        .empty-state svg {
            width: 100px;
            height: 100px;
            margin-bottom: 20px;
            opacity: 0.3;
        }

        /* Modal */
        .modal {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.5);
            backdrop-filter: blur(5px);
            z-index: 1000;
            align-items: center;
            justify-content: center;
        }

        .modal.active {
            display: flex;
        }

        .modal-content {
            background: white;
            padding: 35px;
            border-radius: 20px;
            max-width: 500px;
            width: 90%;
            box-shadow: 0 25px 80px rgba(0,0,0,0.3);
            animation: modalSlideIn 0.3s ease;
        }

        @keyframes modalSlideIn {
            from {
                transform: translateY(-50px);
                opacity: 0;
            }
            to {
                transform: translateY(0);
                opacity: 1;
            }
        }

        .modal-content h2 {
            margin-bottom: 25px;
            color: var(--dark);
            font-family: 'DM Serif Display', serif;
        }

        .modal-form .form-group {
            margin-bottom: 20px;
        }

        .modal-actions {
            display: flex;
            gap: 10px;
            margin-top: 25px;
        }

        .modal-actions button {
            flex: 1;
            padding: 12px;
            border: none;
            border-radius: 10px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
        }

        .btn-cancel {
            background: var(--border);
            color: var(--dark);
        }

        .btn-submit {
            background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%);
            color: white;
        }

        /* Animations */
        @keyframes fadeInDown {
            from {
                opacity: 0;
                transform: translateY(-30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        /* Responsive */
        @media (max-width: 1024px) {
            .main-layout {
                grid-template-columns: 1fr;
            }

            .sidebar {
                position: static;
            }

            .form-grid {
                grid-template-columns: 1fr;
            }
        }

        @media (max-width: 768px) {
            .header h1 {
                font-size: 2.5em;
            }

            .trip-header {
                flex-direction: column;
                align-items: flex-start;
                gap: 15px;
            }

            .trip-actions {
                width: 100%;
                flex-direction: column;
            }

            .btn {
                width: 100%;
                justify-content: center;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <h1>✈️ Trip Expense Manager</h1>
            <p>Track your travel expenses with style</p>
        </div>

        <!-- Main Layout -->
        <div class="main-layout">
            <!-- Sidebar -->
            <div class="sidebar">
                <h2>🗂️ Your Trips</h2>
                <button class="new-trip-btn" onclick="openNewTripModal()">
                    ➕ New Trip
                </button>
                <div class="trip-list" id="tripList">
                    <div class="empty-trips">
                        <p>No trips yet. Create your first trip to get started!</p>
                    </div>
                </div>
            </div>

            <!-- Main Content -->
            <div class="main-content">
                <div id="noTripSelected" class="empty-state">
                    <svg fill="currentColor" viewBox="0 0 20 20">
                        <path d="M10.894 2.553a1 1 0 00-1.788 0l-7 14a1 1 0 001.169 1.409l5-1.429A1 1 0 009 15.571V11a1 1 0 112 0v4.571a1 1 0 00.725.962l5 1.428a1 1 0 001.17-1.408l-7-14z"/>
                    </svg>
                    <h2>Welcome to Trip Expense Manager!</h2>
                    <p>Select a trip or create a new one to start tracking expenses</p>
                </div>

                <div id="tripDetails" style="display: none;">
                    <!-- Trip Header -->
                    <div class="trip-header">
                        <div class="trip-header-left">
                            <h2 id="tripName">Trip Name</h2>
                            <button class="edit-trip-btn" onclick="openEditTripModal()">✏️ Edit</button>
                        </div>
                        <div class="trip-actions">
                            <button class="btn btn-complete" id="completeTripBtn" onclick="toggleTripStatus()">
                                ✓ Complete Trip
                            </button>
                            <button class="btn btn-export" onclick="exportTrip()">
                                📊 Export to Excel
                            </button>
                            <button class="btn btn-export" onclick="exportTripPDF()" style="background: var(--danger);">
                                📄 Export to PDF
                            </button>
                            <button class="btn btn-delete" onclick="deleteTrip()">
                                🗑️ Delete
                            </button>
                        </div>
                    </div>

                    <!-- Budget Overview -->
                    <div class="budget-overview">
                        <div class="budget-card budget">
                            <h3>Total Budget</h3>
                            <div class="amount" id="totalBudget">₹0</div>
                        </div>
                        <div class="budget-card spent">
                            <h3>Total Spent</h3>
                            <div class="amount" id="totalSpent">₹0</div>
                            <div class="small-text" id="expenseCount">0 expenses</div>
                        </div>
                        <div class="budget-card remaining">
                            <h3>Remaining</h3>
                            <div class="amount" id="remaining">₹0</div>
                        </div>
                    </div>

                    <!-- Add Expense Section -->
                    <div class="add-expense-section">
                        <h3>➕ Add New Expense</h3>
                        <div class="form-grid" style="grid-template-columns: 1fr 1fr 1fr 1.5fr;">
                            <div class="form-group">
                                <label>Amount (₹)</label>
                                <input type="number" id="expenseAmount" placeholder="0.00" step="0.01" min="0">
                            </div>
                            <div class="form-group">
                                <label>Category</label>
                                <select id="expenseCategory">
                                    <option value="🍔 Food">🍔 Food</option>
                                    <option value="✈️ Travel">✈️ Travel</option>
                                    <option value="🏨 Accommodation">🏨 Accommodation</option>
                                    <option value="🎭 Entertainment">🎭 Entertainment</option>
                                    <option value="🛍️ Shopping">🛍️ Shopping</option>
                                    <option value="📱 Other">📱 Other</option>
                                </select>
                            </div>
                            <div class="form-group">
                                <label>Person Name</label>
                                <input type="text" id="expensePerson" list="personSuggestions" placeholder="Enter your name" autocomplete="off">
                                <datalist id="personSuggestions"></datalist>
                            </div>
                            <div class="form-group">
                                <label>Description</label>
                                <input type="text" id="expenseDescription" placeholder="What was this for?">
                            </div>
                        </div>
                        
                        <div style="margin-top: 20px; display: grid; grid-template-columns: 1fr auto; gap: 15px; align-items: start;">
                            <div class="form-group">
                                <label>📸 Add Receipt/Photo (Optional)</label>
                                <input type="file" id="expenseImage" accept="image/*" onchange="previewImage(this)">
                                <button class="camera-btn" onclick="capturePhoto()">📷 Take Photo</button>
                                <img id="imagePreview" class="image-preview" alt="Preview">
                                <input type="hidden" id="imageData">
                            </div>
                            <button class="btn-add" onclick="addExpense()" style="margin-top: 28px;">Add Expense</button>
                        </div>
                    </div>

                    <!-- Person-wise Breakdown -->
                    <div id="personBreakdown" style="display: none; margin-bottom: 30px;">
                        <div style="background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%); padding: 20px; border-radius: 15px;">
                            <h3 style="margin-bottom: 15px; color: var(--dark);">👥 Expense by Person</h3>
                            <div id="personBreakdownContent" style="display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 15px;"></div>
                        </div>
                    </div>

                    <!-- Expense Table -->
                    <div class="expense-table-container">
                        <h3>📋 Expense Breakdown</h3>
                        <div id="expenseTableContainer">
                            <div class="empty-state">
                                <p>No expenses yet. Add your first expense above!</p>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- New Trip Modal -->
    <div class="modal" id="newTripModal">
        <div class="modal-content">
            <h2>Create New Trip</h2>
            <div class="modal-form">
                <div class="form-group">
                    <label>Trip Name</label>
                    <input type="text" id="newTripName" placeholder="e.g., Goa Vacation">
                </div>
                <div class="form-group">
                    <label>Budget (₹) <span style="color: #64748b; font-size: 0.85em; font-style: italic;">- Optional</span></label>
                    <input type="number" id="newTripBudget" placeholder="Leave empty if no budget" step="0.01" min="0">
                </div>
                <div class="modal-actions">
                    <button class="btn-cancel" onclick="closeNewTripModal()">Cancel</button>
                    <button class="btn-submit" onclick="createTrip()">Create Trip</button>
                </div>
            </div>
        </div>
    </div>

    <!-- Edit Trip Modal -->
    <div class="modal" id="editTripModal">
        <div class="modal-content">
            <h2>Edit Trip Details</h2>
            <div class="modal-form">
                <div class="form-group">
                    <label>Trip Name</label>
                    <input type="text" id="editTripName" placeholder="e.g., Goa Vacation">
                </div>
                <div class="form-group">
                    <label>Budget (₹) <span style="color: #64748b; font-size: 0.85em; font-style: italic;">- Optional</span></label>
                    <input type="number" id="editTripBudget" placeholder="Leave empty if no budget" step="0.01" min="0">
                </div>
                <div class="modal-actions">
                    <button class="btn-cancel" onclick="closeEditTripModal()">Cancel</button>
                    <button class="btn-submit" onclick="saveEditTrip()">Save Changes</button>
                </div>
            </div>
        </div>
    </div>

    <script>
        const API_URL = '/api';
        let currentTripId = null;
        let currentTrip = null;

        // Image handling
        function previewImage(input) {
            const preview = document.getElementById('imagePreview');
            const imageData = document.getElementById('imageData');
            
            if (input.files && input.files[0]) {
                const reader = new FileReader();
                
                reader.onload = function(e) {
                    preview.src = e.target.result;
                    preview.classList.add('active');
                    imageData.value = e.target.result;
                };
                
                reader.readAsDataURL(input.files[0]);
            }
        }

        async function capturePhoto() {
            try {
                const stream = await navigator.mediaDevices.getUserMedia({ video: true });
                
                // Create video element
                const video = document.createElement('video');
                video.srcObject = stream;
                video.play();
                
                // Create modal for camera
                const modal = document.createElement('div');
                modal.style.cssText = 'position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.9); z-index: 10000; display: flex; flex-direction: column; align-items: center; justify-content: center;';
                
                video.style.cssText = 'max-width: 90%; max-height: 70%; border-radius: 10px;';
                
                const captureBtn = document.createElement('button');
                captureBtn.textContent = '📸 Capture';
                captureBtn.style.cssText = 'margin-top: 20px; padding: 15px 40px; background: #10b981; color: white; border: none; border-radius: 10px; font-size: 1.2em; cursor: pointer;';
                
                const closeBtn = document.createElement('button');
                closeBtn.textContent = '✖ Close';
                closeBtn.style.cssText = 'margin-top: 10px; padding: 10px 30px; background: #ef4444; color: white; border: none; border-radius: 10px; cursor: pointer;';
                
                modal.appendChild(video);
                modal.appendChild(captureBtn);
                modal.appendChild(closeBtn);
                document.body.appendChild(modal);
                
                captureBtn.onclick = function() {
                    const canvas = document.createElement('canvas');
                    canvas.width = video.videoWidth;
                    canvas.height = video.videoHeight;
                    canvas.getContext('2d').drawImage(video, 0, 0);
                    
                    const imageDataUrl = canvas.toDataURL('image/jpeg');
                    document.getElementById('imagePreview').src = imageDataUrl;
                    document.getElementById('imagePreview').classList.add('active');
                    document.getElementById('imageData').value = imageDataUrl;
                    
                    stream.getTracks().forEach(track => track.stop());
                    document.body.removeChild(modal);
                };
                
                closeBtn.onclick = function() {
                    stream.getTracks().forEach(track => track.stop());
                    document.body.removeChild(modal);
                };
                
            } catch (error) {
                alert('Camera access denied or not available');
                console.error('Camera error:', error);
            }
        }

        // Load trips on page load
        window.onload = function() {
            loadTrips();
            loadPersonSuggestions();
        };

        // Person name suggestions
        function loadPersonSuggestions() {
            const savedNames = localStorage.getItem('personNames');
            if (savedNames) {
                const names = JSON.parse(savedNames);
                updatePersonSuggestions(names);
            }
        }

        function updatePersonSuggestions(names) {
            const datalist = document.getElementById('personSuggestions');
            datalist.innerHTML = names.map(name => `<option value="${name}">`).join('');
        }

        function savePersonName(name) {
            if (!name || name.trim() === '') return;
            
            const savedNames = localStorage.getItem('personNames');
            let names = savedNames ? JSON.parse(savedNames) : [];
            
            // Add name if it doesn't exist
            if (!names.includes(name)) {
                names.push(name);
                localStorage.setItem('personNames', JSON.stringify(names));
                updatePersonSuggestions(names);
            }
        }

        // New Trip Modal
        function openNewTripModal() {
            document.getElementById('newTripModal').classList.add('active');
        }

        function closeNewTripModal() {
            document.getElementById('newTripModal').classList.remove('active');
            document.getElementById('newTripName').value = '';
            document.getElementById('newTripBudget').value = '';
        }

        // Edit Trip Modal
        function openEditTripModal() {
            if (!currentTrip) return;
            
            document.getElementById('editTripName').value = currentTrip.name;
            document.getElementById('editTripBudget').value = currentTrip.budget;
            document.getElementById('editTripModal').classList.add('active');
        }

        function closeEditTripModal() {
            document.getElementById('editTripModal').classList.remove('active');
        }

        async function saveEditTrip() {
            const name = document.getElementById('editTripName').value.trim();
            const budgetRaw = document.getElementById('editTripBudget').value.trim();
            const budget = budgetRaw ? parseFloat(budgetRaw) : null;

            if (!name) {
                alert('Please enter a trip name');
                return;
            }

            try {
                await fetch(`${API_URL}/trips/${currentTripId}`, {
                    method: 'PUT',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ name, budget })
                });

                closeEditTripModal();
                loadTrips();
                loadTripSummary(currentTripId);
            } catch (error) {
                console.error('Error updating trip:', error);
                alert('Failed to update trip');
            }
        }

        // Create Trip
        async function createTrip() {
            const name = document.getElementById('newTripName').value.trim();
            const budgetRaw = document.getElementById('newTripBudget').value.trim();
            const budget = budgetRaw ? parseFloat(budgetRaw) : null;

            if (!name) {
                alert('Please enter a trip name');
                return;
            }

            try {
                const response = await fetch(`${API_URL}/trips`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ name, budget })
                });

                const trip = await response.json();
                closeNewTripModal();
                loadTrips();
                selectTrip(trip.id);
            } catch (error) {
                console.error('Error creating trip:', error);
                alert('Failed to create trip. Make sure the backend is running.');
            }
        }

        // Load Trips
        async function loadTrips() {
            try {
                const response = await fetch(`${API_URL}/trips`);
                const trips = await response.json();

                const tripList = document.getElementById('tripList');
                
                if (trips.length === 0) {
                    tripList.innerHTML = '<div class="empty-trips"><p>No trips yet. Create your first trip to get started!</p></div>';
                    return;
                }

                tripList.innerHTML = trips.map(trip => `
                    <div class="trip-card ${trip.id === currentTripId ? 'active' : ''} ${trip.status === 'completed' ? 'completed' : ''}" 
                         onclick="selectTrip('${trip.id}')">
                        <span class="status-badge">${trip.status === 'completed' ? '✓ Done' : '⏳ Ongoing'}</span>
                        <h3>${trip.name}</h3>
                        <div class="budget">Budget: ₹${trip.budget.toLocaleString()}</div>
                    </div>
                `).join('');
            } catch (error) {
                console.error('Error loading trips:', error);
            }
        }

        // Select Trip
        async function selectTrip(tripId) {
            currentTripId = tripId;
            document.getElementById('noTripSelected').style.display = 'none';
            document.getElementById('tripDetails').style.display = 'block';
            
            loadTrips();
            loadTripSummary(tripId);
            loadExpenses(tripId);
        }

        // Load Trip Summary
        async function loadTripSummary(tripId) {
            try {
                const response = await fetch(`${API_URL}/trips/${tripId}/summary`);
                const summary = await response.json();

                currentTrip = summary.trip;

                document.getElementById('tripName').textContent = summary.trip.name;
                document.getElementById('totalBudget').textContent = `₹${summary.total_budget.toLocaleString()}`;
                document.getElementById('totalSpent').textContent = `₹${summary.total_spent.toLocaleString()}`;
                document.getElementById('remaining').textContent = `₹${summary.remaining.toLocaleString()}`;
                document.getElementById('expenseCount').textContent = `${summary.expense_count} expense${summary.expense_count !== 1 ? 's' : ''}`;
                
                // Update complete/reopen button
                const completeBtn = document.getElementById('completeTripBtn');
                if (summary.trip.status === 'completed') {
                    completeBtn.textContent = 'Reopen Trip';
                    completeBtn.classList.add('reopening');
                } else {
                    completeBtn.textContent = '✓ Complete Trip';
                    completeBtn.classList.remove('reopening');
                }
            } catch (error) {
                console.error('Error loading trip summary:', error);
            }
        }

        // Load Expenses
        async function loadExpenses(tripId) {
            try {
                const response = await fetch(`${API_URL}/expenses?trip_id=${tripId}`);
                const expenses = await response.json();

                const container = document.getElementById('expenseTableContainer');

                if (expenses.length === 0) {
                    container.innerHTML = '<div class="empty-state"><p>No expenses yet. Add your first expense above!</p></div>';
                    document.getElementById('personBreakdown').style.display = 'none';
                    return;
                }

                // Calculate person-wise breakdown
                const personTotals = {};
                expenses.forEach(expense => {
                    const person = expense.person || 'Unknown';
                    if (!personTotals[person]) {
                        personTotals[person] = 0;
                    }
                    personTotals[person] += parseFloat(expense.amount);
                });

                // Display person breakdown
                const breakdownContainer = document.getElementById('personBreakdownContent');
                const personBreakdownHTML = Object.entries(personTotals).map(([person, total]) => `
                    <div style="background: white; padding: 15px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                        <div style="font-size: 0.9em; color: #64748b; margin-bottom: 5px;">${person}</div>
                        <div style="font-size: 1.5em; font-weight: 700; color: #667eea;">₹${total.toLocaleString()}</div>
                        <div style="font-size: 0.8em; color: #64748b; margin-top: 5px;">
                            ${expenses.filter(e => (e.person || 'Unknown') === person).length} expense${expenses.filter(e => (e.person || 'Unknown') === person).length !== 1 ? 's' : ''}
                        </div>
                    </div>
                `).join('');
                
                breakdownContainer.innerHTML = personBreakdownHTML;
                document.getElementById('personBreakdown').style.display = 'block';

                let tableHTML = `
                    <table>
                        <thead>
                            <tr>
                                <th>Date</th>
                                <th>Category</th>
                                <th>Amount</th>
                                <th>Person</th>
                                <th>Description</th>
                                <th>Receipt</th>
                                <th>Action</th>
                            </tr>
                        </thead>
                        <tbody>
                `;

                expenses.forEach(expense => {
                    const receiptBtn = expense.image 
                        ? `<button onclick="viewImage('${expense.image}')" style="background: #667eea; color: white; border: none; padding: 5px 10px; border-radius: 5px; cursor: pointer;">📷 View</button>`
                        : '-';
                    
                    tableHTML += `
                        <tr>
                            <td>${expense.date}<br><small style="color: #64748b;">${expense.time}</small></td>
                            <td>${expense.category}</td>
                            <td><strong>₹${parseFloat(expense.amount).toLocaleString()}</strong></td>
                            <td><strong style="color: #667eea;">${expense.person || '-'}</strong></td>
                            <td>${expense.description || '-'}</td>
                            <td>${receiptBtn}</td>
                            <td>
                                <button class="delete-expense-btn" onclick="deleteExpense('${expense.id}')">
                                    Delete
                                </button>
                            </td>
                        </tr>
                    `;
                });

                tableHTML += '</tbody></table>';
                container.innerHTML = tableHTML;
            } catch (error) {
                console.error('Error loading expenses:', error);
            }
        }

        // View uploaded image
        function viewImage(imageData) {
            const modal = document.createElement('div');
            modal.style.cssText = 'position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.9); z-index: 10000; display: flex; align-items: center; justify-content: center; padding: 20px;';
            
            const img = document.createElement('img');
            img.src = imageData;
            img.style.cssText = 'max-width: 90%; max-height: 90%; border-radius: 10px; box-shadow: 0 10px 50px rgba(0,0,0,0.5);';
            
            const closeBtn = document.createElement('button');
            closeBtn.textContent = '✖ Close';
            closeBtn.style.cssText = 'position: absolute; top: 20px; right: 20px; padding: 10px 20px; background: #ef4444; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 1em;';
            
            modal.appendChild(img);
            modal.appendChild(closeBtn);
            document.body.appendChild(modal);
            
            closeBtn.onclick = () => document.body.removeChild(modal);
            modal.onclick = (e) => {
                if (e.target === modal) document.body.removeChild(modal);
            };
        }

        // Export to PDF
        function exportTripPDF() {
            window.open(`${API_URL}/export-pdf/${currentTripId}`, '_blank');
        }

        // Add Expense
        async function addExpense() {
            const amount = parseFloat(document.getElementById('expenseAmount').value);
            const category = document.getElementById('expenseCategory').value;
            const person = document.getElementById('expensePerson').value.trim();
            const description = document.getElementById('expenseDescription').value.trim();
            const imageData = document.getElementById('imageData').value;

            if (!amount || amount <= 0) {
                alert('Please enter a valid amount');
                return;
            }

            if (!person) {
                alert('Please enter your name');
                return;
            }

            try {
                await fetch(`${API_URL}/expenses`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        trip_id: currentTripId,
                        amount,
                        category,
                        person,
                        description,
                        image: imageData
                    })
                });

                // Save person name for future suggestions
                savePersonName(person);

                // Clear form (but keep person name for convenience)
                document.getElementById('expenseAmount').value = '';
                document.getElementById('expenseDescription').value = '';
                document.getElementById('expenseImage').value = '';
                document.getElementById('imageData').value = '';
                document.getElementById('imagePreview').classList.remove('active');

                // Reload data
                loadTripSummary(currentTripId);
                loadExpenses(currentTripId);
            } catch (error) {
                console.error('Error adding expense:', error);
                alert('Failed to add expense');
            }
        }

        // Delete Expense
        async function deleteExpense(expenseId) {
            if (!confirm('Are you sure you want to delete this expense?')) return;

            try {
                await fetch(`${API_URL}/expenses/${expenseId}`, {
                    method: 'DELETE'
                });

                loadTripSummary(currentTripId);
                loadExpenses(currentTripId);
            } catch (error) {
                console.error('Error deleting expense:', error);
            }
        }

        // Toggle Trip Status (Complete/Reopen)
        async function toggleTripStatus() {
            if (!currentTrip) return;

            const isCompleted = currentTrip.status === 'completed';
            const action = isCompleted ? 'reopen' : 'complete';
            const newStatus = isCompleted ? 'ongoing' : 'completed';
            const confirmMessage = isCompleted 
                ? 'Reopen this trip and continue adding expenses?' 
                : 'Mark this trip as completed?';

            if (!confirm(confirmMessage)) return;

            try {
                await fetch(`${API_URL}/trips/${currentTripId}`, {
                    method: 'PUT',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ status: newStatus })
                });

                loadTrips();
                loadTripSummary(currentTripId);
                
                const message = isCompleted 
                    ? 'Trip reopened! You can now add more expenses.' 
                    : 'Trip marked as completed!';
                alert(message);
            } catch (error) {
                console.error('Error toggling trip status:', error);
            }
        }

        // Delete Trip
        async function deleteTrip() {
            if (!confirm('Are you sure you want to delete this trip and all its expenses?')) return;

            try {
                await fetch(`${API_URL}/trips/${currentTripId}`, {
                    method: 'DELETE'
                });

                currentTripId = null;
                currentTrip = null;
                document.getElementById('noTripSelected').style.display = 'block';
                document.getElementById('tripDetails').style.display = 'none';
                loadTrips();
            } catch (error) {
                console.error('Error deleting trip:', error);
            }
        }

        // Export to Excel
        function exportTrip() {
            window.open(`${API_URL}/export/${currentTripId}`, '_blank');
        }
    </script>
</body>
</html>
"""

@app.route('/')
def serve_index():
    return Response(HTML_CONTENT, mimetype='text/html')

# ============================================================
# TRIP ROUTES
# ============================================================
@app.route('/api/trips', methods=['GET'])
def get_trips():
    return jsonify(TRIPS)

@app.route('/api/trips', methods=['POST'])
def create_trip():
    data = request.json
    budget_value = data.get('budget')
    budget = float(budget_value) if budget_value and str(budget_value).strip() else None

    new_trip = {
        'id': str(uuid.uuid4()),
        'name': data.get('name'),
        'budget': budget,
        'status': 'ongoing',
        'created_at': datetime.now().isoformat()
    }
    TRIPS.append(new_trip)
    return jsonify(new_trip), 201

@app.route('/api/trips/<trip_id>', methods=['PUT'])
def update_trip(trip_id):
    data = request.json
    for trip in TRIPS:
        if trip['id'] == trip_id:
            if 'status' in data:
                trip['status'] = data['status']
            if 'name' in data:
                trip['name'] = data['name']
            if 'budget' in data:
                bv = data['budget']
                trip['budget'] = float(bv) if bv and str(bv).strip() else None
            return jsonify(trip)
    return jsonify({'error': 'Trip not found'}), 404

@app.route('/api/trips/<trip_id>', methods=['DELETE'])
def delete_trip(trip_id):
    global TRIPS, EXPENSES
    TRIPS = [t for t in TRIPS if t['id'] != trip_id]
    EXPENSES = [e for e in EXPENSES if e['trip_id'] != trip_id]
    return jsonify({'message': 'Trip deleted'})

# ============================================================
# EXPENSE ROUTES
# ============================================================
@app.route('/api/expenses', methods=['GET'])
def get_expenses():
    trip_id = request.args.get('trip_id')
    if trip_id:
        return jsonify([e for e in EXPENSES if e['trip_id'] == trip_id])
    return jsonify(EXPENSES)

@app.route('/api/expenses', methods=['POST'])
def add_expense():
    data = request.json
    new_expense = {
        'id': str(uuid.uuid4()),
        'trip_id': data.get('trip_id'),
        'category': data.get('category'),
        'amount': float(data.get('amount')),
        'description': data.get('description', ''),
        'person': data.get('person', ''),
        'image': data.get('image', ''),
        'date': datetime.now().strftime('%Y-%m-%d'),
        'time': datetime.now().strftime('%H:%M:%S')
    }
    EXPENSES.append(new_expense)
    return jsonify(new_expense), 201

@app.route('/api/expenses/<expense_id>', methods=['DELETE'])
def delete_expense(expense_id):
    global EXPENSES
    EXPENSES = [e for e in EXPENSES if e['id'] != expense_id]
    return jsonify({'message': 'Expense deleted'})

# ============================================================
# TRIP SUMMARY
# ============================================================
@app.route('/api/trips/<trip_id>/summary', methods=['GET'])
def get_trip_summary(trip_id):
    trip = next((t for t in TRIPS if t['id'] == trip_id), None)
    if not trip:
        return jsonify({'error': 'Trip not found'}), 404

    trip_expenses = [e for e in EXPENSES if e['trip_id'] == trip_id]
    total_spent = sum(float(e['amount']) for e in trip_expenses)
    remaining = (float(trip['budget']) - total_spent) if trip['budget'] is not None else None

    categories = {}
    for e in trip_expenses:
        cat = e['category']
        categories[cat] = categories.get(cat, 0) + float(e['amount'])

    return jsonify({
        'trip': trip,
        'total_budget': float(trip['budget']) if trip['budget'] is not None else None,
        'total_spent': total_spent,
        'remaining': remaining,
        'expense_count': len(trip_expenses),
        'categories': categories
    })

# ============================================================
# EXCEL EXPORT
# ============================================================
@app.route('/api/export/<trip_id>', methods=['GET'])
def export_excel(trip_id):
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        trip = next((t for t in TRIPS if t['id'] == trip_id), None)
        if not trip:
            return jsonify({'error': 'Trip not found'}), 404

        trip_expenses = [e for e in EXPENSES if e['trip_id'] == trip_id]
        if not trip_expenses:
            return jsonify({'error': 'No expenses to export'}), 400

        df = pd.DataFrame(trip_expenses)[['date','time','category','amount','person','description']]
        df.columns = ['Date','Time','Category','Amount','Person','Description']

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Expenses', startrow=4)
            ws = writer.sheets['Expenses']

            ws['A1'] = f'TRIP: {trip["name"]}'
            ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells('A1:F1')
            ws.row_dimensions[1].height = 30

            ws['A2'] = f'Budget: Rs.{float(trip["budget"]):,.2f}' if trip['budget'] else 'Budget: Not Set'
            ws['A2'].font = Font(size=11, bold=True)

            hf = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
            hfont = Font(bold=True, color='FFFFFF', size=11)
            border = Border(left=Side(style='thin',color='e2e8f0'), right=Side(style='thin',color='e2e8f0'),
                           top=Side(style='thin',color='e2e8f0'), bottom=Side(style='thin',color='e2e8f0'))

            for col in range(1, 7):
                cell = ws.cell(row=5, column=col)
                cell.fill = hf
                cell.font = hfont
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            ws.row_dimensions[5].height = 25

            for row in range(6, len(df) + 6):
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col == 4:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.font = Font(bold=True)

            for col, width in zip('ABCDEF', [12,10,20,15,18,40]):
                ws.column_dimensions[col].width = width

            last_row = len(df) + 7
            sf = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
            tb = Border(left=Side(style='medium',color='1e293b'), right=Side(style='medium',color='1e293b'),
                       top=Side(style='medium',color='1e293b'), bottom=Side(style='medium',color='1e293b'))

            ws[f'D{last_row}'] = 'TOTAL SPENT:'
            ws[f'D{last_row}'].font = Font(bold=True, size=11)
            ws[f'D{last_row}'].fill = sf
            ws[f'D{last_row}'].border = tb
            ws[f'D{last_row}'].alignment = Alignment(horizontal='right')
            ws[f'E{last_row}'] = sum(float(e['amount']) for e in trip_expenses)
            ws[f'E{last_row}'].font = Font(bold=True, size=12)
            ws[f'E{last_row}'].number_format = '#,##0.00'
            ws[f'E{last_row}'].fill = sf
            ws[f'E{last_row}'].border = tb
            ws[f'E{last_row}'].alignment = Alignment(horizontal='right')
            ws.row_dimensions[last_row].height = 25

            if trip['budget'] is not None:
                remaining = float(trip['budget']) - sum(float(e['amount']) for e in trip_expenses)
                rc = '10b981' if remaining >= 0 else 'ef4444'
                ws[f'D{last_row+1}'] = 'TRIP BUDGET:'
                ws[f'D{last_row+1}'].font = Font(bold=True, size=11)
                ws[f'D{last_row+1}'].fill = sf
                ws[f'D{last_row+1}'].border = tb
                ws[f'D{last_row+1}'].alignment = Alignment(horizontal='right')
                ws[f'E{last_row+1}'] = float(trip['budget'])
                ws[f'E{last_row+1}'].number_format = '#,##0.00'
                ws[f'E{last_row+1}'].fill = sf
                ws[f'E{last_row+1}'].border = tb
                ws[f'E{last_row+1}'].alignment = Alignment(horizontal='right')
                ws[f'D{last_row+2}'] = 'REMAINING:'
                ws[f'D{last_row+2}'].font = Font(bold=True, size=12, color='FFFFFF')
                ws[f'D{last_row+2}'].fill = PatternFill(start_color=rc, end_color=rc, fill_type='solid')
                ws[f'D{last_row+2}'].border = tb
                ws[f'D{last_row+2}'].alignment = Alignment(horizontal='right')
                ws[f'E{last_row+2}'] = remaining
                ws[f'E{last_row+2}'].font = Font(bold=True, size=13, color='FFFFFF')
                ws[f'E{last_row+2}'].number_format = '#,##0.00'
                ws[f'E{last_row+2}'].fill = PatternFill(start_color=rc, end_color=rc, fill_type='solid')
                ws[f'E{last_row+2}'].border = tb
                ws[f'E{last_row+2}'].alignment = Alignment(horizontal='right')
                ws.row_dimensions[last_row+1].height = 25
                ws.row_dimensions[last_row+2].height = 30

            footer_row = last_row + 4 if trip['budget'] else last_row + 2
            ws[f'A{footer_row}'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
            ws[f'A{footer_row}'].font = Font(size=9, italic=True, color='64748b')

        output.seek(0)
        filename = f"{trip['name'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================
# PDF EXPORT
# ============================================================
@app.route('/api/export-pdf/<trip_id>', methods=['GET'])
def export_pdf(trip_id):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER

        trip = next((t for t in TRIPS if t['id'] == trip_id), None)
        if not trip:
            return jsonify({'error': 'Trip not found'}), 404

        trip_expenses = [e for e in EXPENSES if e['trip_id'] == trip_id]
        if not trip_expenses:
            return jsonify({'error': 'No expenses to export'}), 400

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('T', parent=styles['Heading1'], fontSize=24,
            textColor=colors.HexColor('#667eea'), spaceAfter=30, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements.append(Paragraph(f"Trip: {trip['name']}", title_style))
        elements.append(Spacer(1, 12))

        total_spent = sum(float(e['amount']) for e in trip_expenses)
        budget_data = []
        if trip['budget']:
            budget_data.append(['Trip Budget:', f"Rs.{float(trip['budget']):,.2f}"])
        budget_data.append(['Total Spent:', f"Rs.{total_spent:,.2f}"])
        if trip['budget']:
            budget_data.append(['Remaining:', f"Rs.{float(trip['budget']) - total_spent:,.2f}"])

        bt = Table(budget_data, colWidths=[2*inch, 2*inch])
        bt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('TOPPADDING', (0,0), (-1,-1), 12),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0'))
        ]))
        elements.append(bt)
        elements.append(Spacer(1, 30))

        data = [['Date','Category','Amount','Person','Description']]
        for e in trip_expenses:
            desc = e.get('description') or '-'
            data.append([e['date'], e['category'], f"Rs.{float(e['amount']):,.2f}",
                        e.get('person') or '-', desc[:50] + '...' if len(desc) > 50 else desc])

        table = Table(data, colWidths=[1*inch, 1.2*inch, 1*inch, 1*inch, 2.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('TOPPADDING', (0,1), (-1,-1), 8),
            ('BOTTOMPADDING', (0,1), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        filename = f"{trip['name'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("Trip Expense Tracker Started!")
    app.run(host='0.0.0.0', debug=False, port=5000)
